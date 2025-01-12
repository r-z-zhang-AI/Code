import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


# Markdown表格数据
markdown_table = """
|  类别  | 2020年耕地 | 2020年林地 | 2020年草地 | 2020年水域 | 2020年城乡建设用地 |
| 2023年耕地 | 12350 | 300 | 500 | 0 | 0 |
| 2023年林地 | 10.6 | 434600 | 0 | 0 | 0.4 |
| 2023年草地 | 0 | 0 | 17000.4 | 0 | 0 |
| 2023年水域 | 0 | 9.6 | 0 | 7904 | 0 |
| 2023年城乡建设用地 | 1.8 | 0 | 300 | 0 | 7000 |
"""

# 将Markdown表格转换为列表
lines = markdown_table.strip().split("\n")
header = lines[0].strip("|").split("|")
data = [line.strip("|").split("|") for line in lines[1:]]

# 将数据转换为DataFrame
df = pd.DataFrame(data, columns=header)

# 将数据类型转换为合适的类型（保留1位小数）
for col in df.columns:
    try:
        df[col] = df[col].astype(float).round(1)
    except Exception as e:
        continue

# 将DataFrame保存为Excel文件
df.to_excel("output.xlsx", index=False)

# 使用openpyxl调整列宽和行高
workbook = load_workbook("output.xlsx")
worksheet = workbook.active

# 自动调整列宽
for column_cells in worksheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

# 自动调整行高
for row_cells in worksheet.rows:
    height = max(len(str(cell.value)) for cell in row_cells)
    worksheet.row_dimensions[row_cells[0].row].height = height + 10

workbook.save("./Python/earth_usage/output/20to23.xlsx")
print("OK")